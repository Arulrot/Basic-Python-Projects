import tkinter as tk
from openpyxl import Workbook, load_workbook
import os

def save_to_excel(expression, result):
    file_name = "calculator_results.xlsx"
    
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Expression", "Result"])
    
    ws.append([expression, result])
    wb.save(file_name)

def on_click(button_text):
    if button_text == "=":
        try:
            expression = entry.get()
            result = eval(expression)
            entry.delete(0, tk.END)
            entry.insert(tk.END, str(result))
            save_to_excel(expression, result)
        except Exception as e:
            entry.delete(0, tk.END)
            entry.insert(tk.END, "Error")
    elif button_text == "C":
        entry.delete(0, tk.END)
    else:
        entry.insert(tk.END, button_text)

# GUI setup
root = tk.Tk()
root.title("Simple Calculator")

tk.Label(root, text="Simple Calculator", font=("Arial", 16)).pack()
entry = tk.Entry(root, width=25, font=("Arial", 14))
entry.pack()

buttons = [
    ("7", "8", "9", "/"),
    ("4", "5", "6", "*"),
    ("1", "2", "3", "-"),
    ("C", "0", "=", "+")
]

for row in buttons:
    frame = tk.Frame(root)
    frame.pack()
    for char in row:
        btn = tk.Button(frame, text=char, font=("Arial", 14), width=5, command=lambda ch=char: on_click(ch))
        btn.pack(side=tk.LEFT)

root.mainloop()
